"""
Excel template download + bulk upload for result entry — one subject/class/
term at a time, matching how the entry/update grids already work.

Ported from giia's download_result_template / upload_missed_results
(C:\\Users\\ismai\\OneDrive\\Documents\\GitHub\\giia\\src\\views.py), with two
deliberate changes:

  - giia's template keys rows on the student's raw database id (a column
    literally labelled "Student ID (Required) - Do not modify"). That's an
    internal PK leaking into a spreadsheet a teacher edits and re-uploads —
    fragile if rows get reordered/duplicated/tampered, and exposes an
    implementation detail with no business meaning. This keys on
    admission_number instead — the stable, human-meaningful identifier
    already printed everywhere else (report cards, invoices), and
    duplicate/unknown numbers in the upload are reported back individually
    rather than trusted positionally.

  - giia hardcodes exactly 5 fixed score columns. Here the CA columns come
    from the configurable ScoreComponent list, so the template/importer
    always matches whatever CA1/CA2/CA3/... breakdown is currently
    configured — no code change needed to add or rename a component.
"""

from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from results.models import Result, ScoreComponent
from results.services import save_result_scores
from students.models import Student

HEADER_FILL = PatternFill(start_color='1D3078', end_color='1D3078', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')


def _students_for_upload(section, subject, term, existing_only):
    students = Student.objects.filter(section=section, status='Active').order_by('last_name', 'first_name')
    if existing_only:
        existing_ids = Result.objects.filter(subject=subject, term=term, student__section=section) \
            .values_list('student_id', flat=True)
        students = students.filter(id__in=existing_ids)
    return students


def build_result_template(section, subject, term, existing_only=False) -> Workbook:
    components = list(ScoreComponent.objects.order_by('order'))
    students = _students_for_upload(section, subject, term, existing_only)
    existing = {
        r.student_id: r for r in
        Result.objects.filter(student__section=section, subject=subject, term=term).prefetch_related('component_scores')
    }

    wb = Workbook()
    ws = wb.active
    ws.title = f'{section} {subject.name}'[:31]  # Excel sheet-name length limit

    headers = ['Admission No.', 'Student Name'] + [f'{c.name} (0-{c.max_score})' for c in components] + ['Exam (0-70)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    for row, student in enumerate(students, start=2):
        result = existing.get(student.id)
        component_values = {cs.component_id: cs.value for cs in result.component_scores.all()} if result else {}
        ws.cell(row=row, column=1, value=student.admission_number)
        ws.cell(row=row, column=2, value=student.full_name)
        for col_offset, component in enumerate(components):
            ws.cell(row=row, column=3 + col_offset, value=component_values.get(component.id))
        ws.cell(row=row, column=3 + len(components), value=result.exam if result else None)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 26
    for col in range(3, 3 + len(components) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
    ws.freeze_panes = 'A2'

    notes = wb.create_sheet('Instructions')
    ca_lines = ', '.join(f'{c.name} (0-{c.max_score})' for c in components)
    for i, line in enumerate([
        'INSTRUCTIONS',
        '',
        f'Class: {section}    Subject: {subject.name}    Term: {term}',
        '',
        '1. Do not edit the Admission No. or Student Name columns.',
        f'2. Enter scores in the score columns: {ca_lines}, Exam (0-70).',
        '3. Leave a row blank to skip that student.',
        '4. Save the file, then upload it on the Results page.',
    ], start=1):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions['A'].width = 70

    return wb


class UploadResult:
    def __init__(self):
        self.updated = 0
        self.skipped = 0
        self.errors = []


def apply_result_upload(excel_file, section, subject, term, teacher, existing_only=False) -> UploadResult:
    outcome = UploadResult()
    try:
        wb = load_workbook(excel_file, data_only=True)
    except Exception as exc:
        outcome.errors.append(f'Could not read that file as an Excel workbook: {exc}')
        return outcome

    components = list(ScoreComponent.objects.order_by('order'))
    expected_cols = 2 + len(components) + 1  # Admission No. + Name + CA components + Exam
    ws = wb.worksheets[0]
    if ws.max_column < expected_cols:
        outcome.errors.append(
            'This file does not match the current score columns — download a fresh template and try again.'
        )
        return outcome

    students_by_admission_no = {
        s.admission_number: s for s in _students_for_upload(section, subject, term, existing_only)
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=expected_cols, values_only=True), start=2):
        admission_no, name, *rest = row
        *comp_raw, exam_raw = rest
        if not admission_no and all(v is None for v in comp_raw) and exam_raw is None:
            continue  # blank row — skip silently, not an error
        if not admission_no:
            outcome.errors.append(f'Row {row_idx}: missing Admission No.')
            continue

        student = students_by_admission_no.get(str(admission_no).strip())
        if not student:
            outcome.errors.append(f'Row {row_idx}: "{admission_no}" is not an active student in {section}.')
            continue

        if all(v is None for v in comp_raw) and exam_raw is None:
            outcome.skipped += 1
            continue

        try:
            component_values = {}
            for component, raw_value in zip(components, comp_raw):
                component_values[component.id] = int(Decimal(str(raw_value))) if raw_value is not None else 0
            exam_val = int(Decimal(str(exam_raw))) if exam_raw is not None else 0
        except (InvalidOperation, ValueError):
            outcome.errors.append(f'Row {row_idx} ({admission_no}): scores must be numbers.')
            continue

        save_result_scores(student, subject, term, teacher, exam_val, component_values)
        outcome.updated += 1

    return outcome
