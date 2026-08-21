from django.test import TestCase
from django.urls import reverse

from academics.models import AcademicSession, Section, SchoolClass, Subject, Term
from accounts.models import User
from results.models import GradeBoundary, Result, ScoreComponent
from results.services import class_broadsheet, save_result_scores, student_position
from students.models import Student


def seed_boundaries():
    for grade, lo, hi, point, remark in [
        ('A', 70, 100, 5, 'Excellent'), ('B', 60, 69, 4, 'Very Good'), ('C', 50, 59, 3, 'Good'),
        ('D', 45, 49, 2, 'Fair'), ('E', 40, 44, 1, 'Pass'), ('F', 0, 39, 0, 'Fail'),
    ]:
        GradeBoundary.objects.create(grade=grade, min_score=lo, max_score=hi, point=point, remark=remark)


def seed_components():
    """CA1/CA2/CA3, 10 marks each — matches the real seed_score_components
    defaults so tests exercise the same shape production starts with."""
    ScoreComponent.objects.create(name='CA1', max_score=10, order=1)
    ScoreComponent.objects.create(name='CA2', max_score=10, order=2)
    ScoreComponent.objects.create(name='CA3', max_score=10, order=3)
    return list(ScoreComponent.objects.order_by('order'))


def spread_ca(components, total):
    """Distribute `total` marks across components (in given order), each
    capped at its own max_score — lets fixtures say "this student's CA
    total is 25" without caring which component holds how much, mirroring
    the old single-`ca`-field tests' intent."""
    values, remaining = {}, total
    for c in components:
        take = min(remaining, c.max_score)
        values[c.id] = take
        remaining -= take
    return values


class ResultAutoCalcTests(TestCase):
    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        self.section = Section.objects.create(school_class=school_class, name='A')
        self.subject = Subject.objects.create(name='Mathematics', level='Primary')
        self.student = Student.objects.create(
            first_name='Test', last_name='Student', gender='Male', school_class=school_class, section=self.section,
        )

    def test_total_and_grade_computed_from_components_and_exam(self):
        component_values = {c.id: c.max_score for c in self.components}  # 10+10+10 = 30
        result = save_result_scores(self.student, self.subject, self.term, None, 60, component_values)
        self.assertEqual(result.total, 90)
        self.assertEqual(result.grade, 'A')
        self.assertEqual(result.remark, 'Excellent')

    def test_component_scores_are_clamped_to_their_own_max(self):
        component_values = {c.id: 999 for c in self.components}
        result = save_result_scores(self.student, self.subject, self.term, None, 999, component_values)
        for cs in result.component_scores.all():
            self.assertEqual(cs.value, cs.component.max_score)
        self.assertEqual(result.exam, 70)  # exam still clamps independently, unchanged behaviour
        self.assertEqual(result.total, 30 + 70)

    def test_failing_boundary(self):
        component_values = {c.id: 2 for c in self.components}  # 6
        result = save_result_scores(self.student, self.subject, self.term, None, 9, component_values)
        self.assertEqual(result.total, 15)
        self.assertEqual(result.grade, 'F')


class ReportCardAccessTests(TestCase):
    def setUp(self):
        seed_boundaries()
        session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        section = Section.objects.create(school_class=school_class, name='A')

        self.owner = User.objects.create_user(username='stu1', password='pw', role='student')
        self.other = User.objects.create_user(username='stu2', password='pw', role='student')
        self.student = Student.objects.create(
            first_name='Owner', last_name='Student', gender='Male', school_class=school_class,
            section=section, user=self.owner,
        )

    def test_student_can_view_own_report_card(self):
        self.client.force_login(self.owner)
        response = self.client.get(reverse('results:report_card', args=[self.student.pk, self.term.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Owner Student')

    def test_student_cannot_view_someone_elses_report_card(self):
        self.client.force_login(self.other)
        response = self.client.get(reverse('results:report_card', args=[self.student.pk, self.term.pk]))
        self.assertContains(response, 'Access denied')


class ClassBroadsheetTests(TestCase):
    """
    Position ranking — adapted from giia's per-subject Result.calculate_position()
    into a whole-term, average-based "Position in Class" (the conventional
    report-card metric), with competition ranking for ties (1, 2, 2, 4 — not 1, 2, 2, 3).
    """

    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        self.section = Section.objects.create(school_class=school_class, name='A')
        self.maths = Subject.objects.create(name='Mathematics', level='Primary')
        self.english = Subject.objects.create(name='English Studies', level='Primary')

        self.top = Student.objects.create(first_name='Top', last_name='Student', gender='Male', school_class=school_class, section=self.section)
        self.tied_a = Student.objects.create(first_name='TiedA', last_name='Student', gender='Female', school_class=school_class, section=self.section)
        self.tied_b = Student.objects.create(first_name='TiedB', last_name='Student', gender='Male', school_class=school_class, section=self.section)
        self.last = Student.objects.create(first_name='Last', last_name='Student', gender='Female', school_class=school_class, section=self.section)
        self.no_results = Student.objects.create(first_name='NoResults', last_name='Student', gender='Male', school_class=school_class, section=self.section)

        save_result_scores(self.top, self.maths, self.term, None, 68, spread_ca(self.components, 30))    # 98
        save_result_scores(self.tied_a, self.maths, self.term, None, 45, spread_ca(self.components, 25))  # 70
        save_result_scores(self.tied_b, self.maths, self.term, None, 50, spread_ca(self.components, 20))  # 70
        save_result_scores(self.last, self.maths, self.term, None, 10, spread_ca(self.components, 5))     # 15

    def test_ties_share_a_position_and_the_next_rank_skips(self):
        board = class_broadsheet(self.section, self.term)
        by_student = {row['student'].id: row for row in board['rows']}

        self.assertEqual(by_student[self.top.id]['position'], 1)
        self.assertEqual(by_student[self.tied_a.id]['position'], 2)
        self.assertEqual(by_student[self.tied_b.id]['position'], 2)
        self.assertEqual(by_student[self.last.id]['position'], 4)  # skips 3 — two students already hold rank 2

    def test_student_with_no_results_is_unranked_not_last(self):
        board = class_broadsheet(self.section, self.term)
        by_student = {row['student'].id: row for row in board['rows']}
        self.assertIsNone(by_student[self.no_results.id]['position'])

    def test_class_size_counts_every_active_student(self):
        board = class_broadsheet(self.section, self.term)
        self.assertEqual(board['class_size'], 5)

    def test_student_position_matches_the_broadsheet(self):
        position, class_size = student_position(self.top, self.term)
        self.assertEqual(position, 1)
        self.assertEqual(class_size, 5)

    def test_report_card_shows_position(self):
        self.client.force_login(User.objects.create_user(username='admin1', password='pw', role='admin'))
        response = self.client.get(reverse('results:report_card', args=[self.top.pk, self.term.pk]))
        self.assertContains(response, '1 of 5')


class ReportCardPdfTests(TestCase):
    """Ported from giia's download_single_result_pdf (WeasyPrint rendering
    the same HTML template — no separate PDF template)."""

    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        self.section = Section.objects.create(school_class=school_class, name='A')
        subject = Subject.objects.create(name='Mathematics', level='Primary')

        self.owner = User.objects.create_user(username='stu1', password='pw', role='student')
        self.other = User.objects.create_user(username='stu2', password='pw', role='student')
        self.admin = User.objects.create_user(username='admin1', password='pw', role='admin')
        self.student = Student.objects.create(
            first_name='Owner', last_name='Student', gender='Male', school_class=school_class,
            section=self.section, user=self.owner,
        )
        save_result_scores(self.student, subject, self.term, None, 60, spread_ca(self.components, 25))

    def test_admin_can_download_a_students_report_card_as_pdf(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('results:report_card_pdf', args=[self.student.pk, self.term.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertGreater(len(response.content), 500)

    def test_student_cannot_download_someone_elses_report_card_pdf(self):
        self.client.force_login(self.other)
        response = self.client.get(reverse('results:report_card_pdf', args=[self.student.pk, self.term.pk]))
        self.assertRedirects(response, reverse('portal:home'))


class BulkReportCardPdfTests(TestCase):
    """Deliberate deviation from giia: giia's "download all" is a ZIP of one
    PDF per student. This renders a single combined PDF instead — one
    page-section per student — per the explicit spec (see results/pdf.py)."""

    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        self.session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=self.session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        self.section = Section.objects.create(school_class=school_class, name='A')
        self.subject = Subject.objects.create(name='Mathematics', level='Primary')
        self.admin = User.objects.create_user(username='admin1', password='pw', role='admin')

        self.s1 = Student.objects.create(first_name='One', last_name='Student', gender='Male', school_class=school_class, section=self.section)
        self.s2 = Student.objects.create(first_name='Two', last_name='Student', gender='Female', school_class=school_class, section=self.section)
        save_result_scores(self.s1, self.subject, self.term, None, 60, spread_ca(self.components, 25))
        save_result_scores(self.s2, self.subject, self.term, None, 55, spread_ca(self.components, 20))

    def test_class_results_pdf_is_a_single_pdf_with_one_page_section_per_student(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('results:class_results_pdf'), {'section_id': self.section.pk, 'term_id': self.term.pk})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))

        from io import BytesIO
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(response.content))
        self.assertGreaterEqual(len(reader.pages), 2)  # both active students present, not zero/one

    def test_bulk_pdf_respects_a_non_current_term_id(self):
        past_session = AcademicSession.objects.create(name='2024/2025', is_current=False)
        past_term = Term.objects.create(session=past_session, name='third', is_current=False)
        save_result_scores(self.s1, self.subject, past_term, None, 40, spread_ca(self.components, 15))

        self.client.force_login(self.admin)
        response = self.client.get(reverse('results:class_results_pdf'), {'section_id': self.section.pk, 'term_id': past_term.pk})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')


class ResultExcelTests(TestCase):
    """Ported from giia's download_result_template / upload_missed_results,
    matched on admission_number instead of raw student PK (see results/excel.py),
    with CA columns generated from the configured ScoreComponents instead of
    a fixed layout."""

    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        self.section = Section.objects.create(school_class=school_class, name='A')
        self.subject = Subject.objects.create(name='Mathematics', level='Primary')
        self.student = Student.objects.create(
            first_name='Test', last_name='Student', gender='Male', school_class=school_class, section=self.section,
        )
        self.admin = User.objects.create_user(username='admin1', password='pw', role='admin')

    def test_template_download_contains_the_students_admission_number_and_dynamic_ca_columns(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('results:download_template'), {
            'section_id': self.section.pk, 'subject_id': self.subject.pk, 'term_id': self.term.pk,
        })
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])

        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, 'Admission No.')
        self.assertEqual(ws.cell(row=2, column=1).value, self.student.admission_number)
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn('CA1 (0-10)', header_row)
        self.assertIn('CA2 (0-10)', header_row)
        self.assertIn('CA3 (0-10)', header_row)
        self.assertIn('Exam (0-70)', header_row)

    def test_upload_applies_scores_matched_by_admission_number(self):
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['Admission No.', 'Student Name', 'CA1 (0-10)', 'CA2 (0-10)', 'CA3 (0-10)', 'Exam (0-70)'])
        ws.append([self.student.admission_number, self.student.full_name, 9, 9, 10, 65])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'upload.xlsx'

        self.client.force_login(self.admin)
        response = self.client.post(reverse('results:upload'), {
            'section_id': self.section.pk, 'subject_id': self.subject.pk, 'term_id': self.term.pk, 'excel_file': buffer,
        })
        self.assertEqual(response.status_code, 302)

        result = Result.objects.get(student=self.student, subject=self.subject, term=self.term)
        self.assertEqual(result.exam, 65)
        self.assertEqual(result.total, 9 + 9 + 10 + 65)

    def test_upload_reports_an_unknown_admission_number_without_crashing(self):
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['Admission No.', 'Student Name', 'CA1 (0-10)', 'CA2 (0-10)', 'CA3 (0-10)', 'Exam (0-70)'])
        ws.append(['GFA/DOES/NOT/EXIST', 'Ghost Student', 5, 5, 5, 50])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'upload.xlsx'

        self.client.force_login(self.admin)
        response = self.client.post(reverse('results:upload'), {
            'section_id': self.section.pk, 'subject_id': self.subject.pk, 'term_id': self.term.pk, 'excel_file': buffer,
        }, follow=True)
        self.assertContains(response, 'not an active student')
        self.assertEqual(Result.objects.count(), 0)


class ScoreComponentConfigurabilityTests(TestCase):
    """Proves CA components are genuinely configurable — adding a 4th one
    (no code change, just a new ScoreComponent row) flows through the entry
    grid, into the saved Result, and onto the report card."""

    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        self.session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=self.session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        self.section = Section.objects.create(school_class=school_class, name='A')
        self.subject = Subject.objects.create(name='Mathematics', level='Primary')
        self.student = Student.objects.create(
            first_name='Test', last_name='Student', gender='Male', school_class=school_class, section=self.section,
        )
        self.admin = User.objects.create_user(username='admin1', password='pw', role='admin')
        self.client.force_login(self.admin)

    def test_adding_a_fourth_component_flows_through_entry_grid_to_report_card(self):
        ca4 = ScoreComponent.objects.create(name='CA4', max_score=10, order=4)

        post_data = {
            'session_id': self.session.pk, 'term_id': self.term.pk,
            'section_id': self.section.pk, 'subject_id': self.subject.pk,
            f'exam_{self.student.pk}': 40,
        }
        for c in self.components:
            post_data[f'comp_{c.id}_{self.student.pk}'] = 5
        post_data[f'comp_{ca4.id}_{self.student.pk}'] = 8
        response = self.client.post(reverse('results:entry'), post_data)
        self.assertEqual(response.status_code, 200)

        result = Result.objects.get(student=self.student, subject=self.subject, term=self.term)
        self.assertEqual(result.total, 5 + 5 + 5 + 8 + 40)  # 63

        response = self.client.get(reverse('results:report_card', args=[self.student.pk, self.term.pk]))
        self.assertContains(response, 'CA4')


class HistoricalReportCardTests(TestCase):
    """A past (non-current) session/term's results must stay reachable —
    nothing here is restricted to "current term"."""

    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        self.current_session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        Term.objects.create(session=self.current_session, name='first', is_current=True)
        self.past_session = AcademicSession.objects.create(name='2024/2025', is_current=False)
        self.past_term = Term.objects.create(session=self.past_session, name='third', is_current=False)

        school_class = SchoolClass.objects.create(name='JSS 2', level='Secondary', order=8)
        self.section = Section.objects.create(school_class=school_class, name='A')
        self.subject = Subject.objects.create(name='English Language', level='Secondary')
        self.student = Student.objects.create(
            first_name='Past', last_name='Student', gender='Male', school_class=school_class, section=self.section,
        )
        save_result_scores(self.student, self.subject, self.past_term, None, 55, spread_ca(self.components, 22))

        self.admin = User.objects.create_user(username='admin1', password='pw', role='admin')
        self.client.force_login(self.admin)

    def test_report_card_reachable_for_a_past_term(self):
        response = self.client.get(reverse('results:report_card', args=[self.student.pk, self.past_term.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Past Student')

    def test_report_cards_roster_reachable_for_a_past_session_and_term(self):
        response = self.client.get(reverse('results:report_cards'), {
            'session_id': self.past_session.pk, 'term_id': self.past_term.pk, 'section_id': self.section.pk,
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.student.admission_number)

    def test_class_results_pdf_reachable_for_a_past_term(self):
        response = self.client.get(reverse('results:class_results_pdf'), {
            'section_id': self.section.pk, 'term_id': self.past_term.pk,
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')


class EnterVsUpdateStudentScopeTests(TestCase):
    """Enter Results shows every active student (for first-time entry);
    Update Results shows only students who already have a result (for
    correction) — the core distinction the spec asked for."""

    def setUp(self):
        seed_boundaries()
        self.components = seed_components()
        self.session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        self.term = Term.objects.create(session=self.session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        self.section = Section.objects.create(school_class=school_class, name='A')
        self.subject = Subject.objects.create(name='Mathematics', level='Primary')

        self.has_result = Student.objects.create(first_name='Has', last_name='Result', gender='Male', school_class=school_class, section=self.section)
        self.no_result_a = Student.objects.create(first_name='NoResultA', last_name='Student', gender='Female', school_class=school_class, section=self.section)
        self.no_result_b = Student.objects.create(first_name='NoResultB', last_name='Student', gender='Male', school_class=school_class, section=self.section)
        save_result_scores(self.has_result, self.subject, self.term, None, 40, spread_ca(self.components, 15))

        self.admin = User.objects.create_user(username='admin1', password='pw', role='admin')
        self.client.force_login(self.admin)

    def _params(self):
        return {'session_id': self.session.pk, 'term_id': self.term.pk, 'section_id': self.section.pk, 'subject_id': self.subject.pk}

    def test_enter_results_shows_every_active_student(self):
        response = self.client.get(reverse('results:entry'), self._params())
        for student in [self.has_result, self.no_result_a, self.no_result_b]:
            self.assertContains(response, student.admission_number)

    def test_update_results_shows_only_students_with_an_existing_result(self):
        response = self.client.get(reverse('results:update'), self._params())
        self.assertContains(response, self.has_result.admission_number)
        self.assertNotContains(response, self.no_result_a.admission_number)
        self.assertNotContains(response, self.no_result_b.admission_number)


class BlankQueryParamRobustnessTests(TestCase):
    """A blank or garbage section_id/subject_id (e.g. a <select> with no
    option chosen yet, or a hand-edited URL) must fail gracefully — not
    raise ValueError: Field 'id' expected a number but got ''."""

    def setUp(self):
        seed_boundaries()
        seed_components()
        session = AcademicSession.objects.create(name='2025/2026', is_current=True)
        Term.objects.create(session=session, name='first', is_current=True)
        school_class = SchoolClass.objects.create(name='Primary 5', level='Primary', order=1)
        Section.objects.create(school_class=school_class, name='A')
        self.admin = User.objects.create_user(username='admin1', password='pw', role='admin')
        self.client.force_login(self.admin)

    def test_class_results_pdf_with_blank_section_id_404s_not_500s(self):
        response = self.client.get(reverse('results:class_results_pdf'), {'section_id': ''})
        self.assertEqual(response.status_code, 404)

    def test_class_results_pdf_with_garbage_section_id_404s_not_500s(self):
        response = self.client.get(reverse('results:class_results_pdf'), {'section_id': 'not-a-number'})
        self.assertEqual(response.status_code, 404)

    def test_download_template_with_blank_ids_redirects_not_500s(self):
        response = self.client.get(reverse('results:download_template'), {'section_id': '', 'subject_id': ''})
        self.assertRedirects(response, reverse('results:entry'))

    def test_broadsheet_with_garbage_section_id_falls_back_gracefully(self):
        response = self.client.get(reverse('results:broadsheet'), {'section_id': 'not-a-number'})
        self.assertEqual(response.status_code, 200)
